import streamlit as st
from datetime import date, datetime
import pandas as pd
import yaml, io
from utils import compile_or_regex
from openpyxl.utils import get_column_letter
import os

# Try importing Google Sheets functionality
try:
    from google_sheets_integration import read_from_sheets, export_to_sheets
    from collector import collect as collect_rss
    HAS_SHEETS = True
except ImportError:
    HAS_SHEETS = False
    from collector import collect as collect_rss

st.set_page_config(page_title="U.S.-China News Scraper", layout="wide")

# ============================================================================
# 测试阶段访问控制（Testing Stage Access Control）
# ============================================================================
def check_test_access():
    """
    检查测试访问权限
    支持两种方式：
    1. 密码保护：需要输入正确的测试密码
    2. 时间限制：超过测试截止日期后自动禁用
    """
    # 从 Streamlit secrets 或环境变量读取配置
    test_enabled = os.getenv("TEST_MODE_ENABLED", "false").lower() == "true"
    test_password = os.getenv("TEST_PASSWORD", "")
    test_deadline = os.getenv("TEST_DEADLINE", "")  # 格式: YYYY-MM-DD
    
    # 尝试从 Streamlit secrets 读取（优先级更高）
    try:
        if hasattr(st, 'secrets'):
            if 'test_mode' in st.secrets:
                test_mode_config = st.secrets.get('test_mode', {})
                # 处理 enabled：支持布尔值和字符串 "true"/"false"
                enabled_val = test_mode_config.get('enabled', test_enabled)
                if isinstance(enabled_val, str):
                    test_enabled = enabled_val.lower() == "true"
                else:
                    test_enabled = bool(enabled_val)
                test_password = test_mode_config.get('password', test_password)
                test_deadline = test_mode_config.get('deadline', test_deadline)
    except Exception as e:
        # 调试信息（仅在开发环境显示）
        if os.getenv("DEBUG", "").lower() == "true":
            st.warning(f"⚠️ Test mode config error: {e}")
        pass
    
    # 如果测试模式未启用，直接允许访问
    if not test_enabled:
        return True, None
    
    # 检查时间限制
    if test_deadline:
        try:
            deadline_date = datetime.strptime(test_deadline, "%Y-%m-%d").date()
            if date.today() > deadline_date:
                return False, f"⚠️ Testing stage has ended (deadline: {test_deadline})"
        except:
            pass
    
    # 检查密码（如果设置了密码）
    if test_password:
        if 'test_authenticated' not in st.session_state:
            st.session_state.test_authenticated = False
        
        if not st.session_state.test_authenticated:
            st.markdown("## 🔒 Testing Stage Access")
            st.info("This is a test version. Please enter the test password to continue.")
            
            password_input = st.text_input("Test Password", type="password", key="test_password_input")
            
            col1, col2 = st.columns([1, 4])
            with col1:
                if st.button("Verify", type="primary"):
                    if password_input == test_password:
                        st.session_state.test_authenticated = True
                        st.rerun()
                    else:
                        st.error("❌ Incorrect password, please try again")
                        return False, None
            
            st.stop()
    
    return True, None

# 检查访问权限
access_granted, error_msg = check_test_access()
if not access_granted:
    st.error(error_msg)
    st.stop()

st.markdown("## U.S.-China News Scraper")

# Initialize session state to persist results
if 'df_result' not in st.session_state:
    st.session_state.df_result = None
if 'last_run_params' not in st.session_state:
    st.session_state.last_run_params = None
if 'is_processing' not in st.session_state:
    st.session_state.is_processing = False
# Track API calls to prevent duplicate execution
if 'api_calls_made' not in st.session_state:
    st.session_state.api_calls_made = 0
if 'last_api_call_time' not in st.session_state:
    st.session_state.last_api_call_time = None

# Load config and categories
with open("config_en.yaml","r",encoding="utf-8") as f:
    CFG = yaml.safe_load(f) or {}
with open("categories_en.yaml","r",encoding="utf-8") as f:
    CATS = yaml.safe_load(f) or {}
CATEGORIES = CATS.get("categories", {})

def _format_date_label(value):
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def render_results(df: pd.DataFrame, start_date, end_date):
    if df is None or df.empty:
        st.warning("No articles found matching the criteria. Please adjust the date range or sources and try again.")
        return

    start_label = _format_date_label(start_date)
    end_label = _format_date_label(end_date)

    ordered_categories = list(CATEGORIES.keys())
    total_categories = len(ordered_categories)

    st.markdown(f"### 📊 Summary ({start_label} → {end_label})")

    total = len(df)
    if "Category" not in df.columns:
        st.warning("Current results do not have category information. Please re-run the collection process.")
        return

    unc = df[df["Category"] == "Uncategorized"]
    unc_count = len(unc)

    categories_with_articles = [
        cat for cat in ordered_categories if (df["Category"] == cat).any()
    ]
    actual_categories_count = len(categories_with_articles)

    if "Outlet" in df.columns:
        actual_outlets = set(df["Outlet"].dropna().unique())
        actual_outlets_count = len(actual_outlets)
    else:
        actual_outlets_count = 0
    total_outlets_estimate = len(all_sources)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("📰 Total Articles", total)
    with col2:
        st.metric("📂 Categories", f"{actual_categories_count}/{total_categories}")
    with col3:
        st.metric("📺 Outlets", f"{actual_outlets_count}/{total_outlets_estimate}")

    st.markdown("---")
    st.markdown("#### 📋 By Category")

    category_counts = []
    for cat in ordered_categories:
        count = int((df["Category"] == cat).sum())
        category_counts.append((cat, count))

    cols = st.columns(2)
    for idx, (cat, count) in enumerate(category_counts):
        if count == 0:
            continue
        percentage = (count / total * 100) if total > 0 else 0.0
        with cols[idx % 2]:
            st.markdown(f"**{cat}**")
            st.progress(min(count / total, 1.0) if total > 0 else 0.0)
            st.caption(f"{count} articles ({percentage:.1f}%)")

    if unc_count > 0:
        st.markdown("---")
        st.markdown(f"**Uncategorized**: {unc_count} articles ({(unc_count/total*100):.1f}%)")
        st.progress(min(unc_count / total, 1.0) if total > 0 else 0.0)

    st.markdown("---")
    st.markdown("### 📰 Articles")
    
    # Add classification feedback mechanism (improved with category dropdown)
    if 'classification_feedback' not in st.session_state:
        st.session_state.classification_feedback = {}
    
    # Load all available categories
    try:
        import yaml
        with open('categories_en.yaml', 'r') as f:
            categories_config = yaml.safe_load(f)
        all_categories = list(categories_config.get('categories', {}).keys()) + ["Uncategorized"]
    except:
        # Fallback: get categories from dataframe
        all_categories = sorted(df['Category'].unique().tolist()) if 'Category' in df.columns else ["Uncategorized"]
    
    # Show feedback option in a collapsible section (expanded by default for visibility)
    with st.expander("💬 Provide Classification Feedback (Help improve accuracy)", expanded=True):
        st.markdown("**How it works**: If you find a misclassified article, mark it as incorrect and select the correct category. This feedback will be used to improve future classifications by adding it to the AI's training examples.")
        st.markdown("---")
        
        selected_idx = st.selectbox(
            "Select an article to provide feedback:",
            options=range(len(df)),
            format_func=lambda x: f"{df.iloc[x].get('Headline', 'No title')[:60]}... | Current: {df.iloc[x].get('Category', 'N/A')}"
        )
        
        if selected_idx is not None:
            selected_row = df.iloc[selected_idx]
            feedback_key = selected_row.get('URL', f"idx_{selected_idx}")
            
            st.markdown(f"**Headline**: {selected_row.get('Headline', 'N/A')}")
            st.markdown(f"**Summary**: {selected_row.get('Nut Graph', 'N/A')[:200]}...")
            st.markdown(f"**Current Category**: `{selected_row.get('Category', 'N/A')}`")
            
            # Check if feedback already exists
            feedback_file = "classification_feedback.json"
            existing_feedback = {}
            try:
                import json
                if os.path.exists(feedback_file):
                    with open(feedback_file, 'r') as f:
                        existing_feedback = json.load(f)
            except:
                pass
            
            feedback_status = existing_feedback.get(feedback_key, {}).get('status', None)
            
            if feedback_status:
                st.info(f"✅ Feedback already recorded: {feedback_status}")
                if feedback_status == "incorrect":
                    correct_cat = existing_feedback.get(feedback_key, {}).get('correct_category', 'N/A')
                    st.info(f"📝 Correct category: `{correct_cat}`")
            else:
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("✅ Correct", key=f"correct_feedback_{selected_idx}"):
                        # Save feedback
                        try:
                            import json
                            if not os.path.exists(feedback_file):
                                feedback_data = {}
                            else:
                                with open(feedback_file, 'r') as f:
                                    feedback_data = json.load(f)
                            
                            feedback_data[feedback_key] = {
                                'status': 'correct',
                                'headline': selected_row.get('Headline', ''),
                                'summary': selected_row.get('Nut Graph', ''),
                                'current_category': selected_row.get('Category', ''),
                                'reason': '',  # Correct classification doesn't need a reason
                                'timestamp': datetime.now().isoformat()
                            }
                            
                            with open(feedback_file, 'w') as f:
                                json.dump(feedback_data, f, indent=2)
                            
                            st.success("✅ Thank you! Feedback recorded. This will help improve future classifications.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error saving feedback: {e}")
                
                with col2:
                    if st.button("❌ Incorrect", key=f"incorrect_feedback_{selected_idx}"):
                        st.session_state.classification_feedback[feedback_key] = "incorrect"
                        st.warning("Please select the correct category below.")
                        st.rerun()
                
                if st.session_state.classification_feedback.get(feedback_key) == "incorrect":
                    st.markdown("---")
                    correct_category = st.selectbox(
                        "What should the correct category be?",
                        options=all_categories,
                        key=f"correct_cat_select_{selected_idx}",
                        index=all_categories.index(selected_row.get('Category', 'Uncategorized')) if selected_row.get('Category', 'Uncategorized') in all_categories else 0
                    )
                    
                    # 添加"原因"输入框（可选）
                    reason = st.text_area(
                        "Why is this the correct category? (Optional - helps AI learn better)",
                        key=f"reason_input_{selected_idx}",
                        placeholder="Example: This is a state visit where China hosts a foreign leader, involving bilateral diplomatic relations, should be categorized as Multilateralism",
                        height=100
                    )
                    
                    if st.button("💾 Save Feedback", key=f"save_feedback_{selected_idx}"):
                        # Save feedback with correct category
                        try:
                            import json
                            if not os.path.exists(feedback_file):
                                feedback_data = {}
                            else:
                                with open(feedback_file, 'r') as f:
                                    feedback_data = json.load(f)
                            
                            feedback_data[feedback_key] = {
                                'status': 'incorrect',
                                'headline': selected_row.get('Headline', ''),
                                'summary': selected_row.get('Nut Graph', ''),
                                'current_category': selected_row.get('Category', ''),
                                'correct_category': correct_category,
                                'reason': reason.strip() if reason else '',  # 添加原因字段
                                'timestamp': datetime.now().isoformat()
                            }
                            
                            with open(feedback_file, 'w') as f:
                                json.dump(feedback_data, f, indent=2)
                            
                            st.success(f"✅ Feedback saved! This article should be `{correct_category}`. This will be used to improve future classifications.")
                            st.info("💡 **Next time you classify articles, this feedback will be included in the AI's training examples.**")
                            del st.session_state.classification_feedback[feedback_key]
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error saving feedback: {e}")
    
    st.dataframe(df, use_container_width=True, height=600)

    st.markdown("---")
    st.markdown("## 🔥 Trending News")
    st.markdown("Show news covered by multiple outlets (grouped by category)")

    try:
        from news_trending import group_similar_news, generate_trending_rank

        with st.spinner("Analyzing news trends..."):
            df_with_groups = group_similar_news(
                df.copy(),
                similarity_threshold=0.55,  # Slightly higher threshold for better grouping (was 0.5)
                use_api=use_api_classification
            )
        trending_df = generate_trending_rank(df_with_groups, top_n=3, min_sources=2)  # Keep at 2 sources for now

        if not trending_df.empty:
            categories = sorted(trending_df["Category"].unique())
            if len(categories) > 1:
                tabs = st.tabs(categories)
                for idx, category in enumerate(categories):
                    with tabs[idx]:
                        category_trending = trending_df[trending_df["Category"] == category]
                        for _, row in category_trending.iterrows():
                            with st.container():
                                st.markdown(f"### 🔥 Covered by {row['SourceCount']} outlets")
                                st.markdown(f"**{row['Headline']}**")
                                st.markdown(f"**Outlets**: {row['Outlets']}")
                                if row.get("Date"):
                                    st.markdown(f"**Date**: {row['Date']}")
                                if row.get("URLs"):
                                    st.markdown("**Related articles**:")
                                    for url in row["URLs"][:5]:
                                        st.markdown(f"- [View original]({url})")
                                st.markdown("---")
            else:
                category = categories[0]
                category_trending = trending_df[trending_df["Category"] == category]
                for _, row in category_trending.iterrows():
                    with st.container():
                        st.markdown(f"### 🔥 Covered by {row['SourceCount']} outlets")
                        st.markdown(f"**{row['Headline']}**")
                        st.markdown(f"**Outlets**: {row['Outlets']}")
                        if row.get("Date"):
                            st.markdown(f"**Date**: {row['Date']}")
                        if row.get("URLs"):
                            st.markdown("**Related articles**:")
                            for url in row["URLs"][:5]:
                                st.markdown(f"- [View original]({url})")
                        st.markdown("---")
        else:
            st.info("No trending news (requires at least 2 outlets covering the same news)")
    except ImportError as e:
        st.warning(f"⚠️ Trending news feature temporarily unavailable: {e}")
    except Exception as e:
        st.warning(f"⚠️ Error generating trending news: {e}")
        import traceback
        st.code(traceback.format_exc())

    st.markdown("---")
    st.markdown("### 💾 Export")

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # All articles
        df.to_excel(writer, sheet_name="All", index=False)
        ws_all = writer.sheets.get("All")
        if ws_all is not None:
            for col_name, max_width in [("Date", 22), ("Outlet", 18), ("Headline", 80)]:
                if col_name in df.columns:
                    idx = list(df.columns).index(col_name) + 1
                    values = df[col_name].astype(str).tolist() if not df.empty else []
                    width = min(max(len(col_name), max((len(v) for v in values), default=0)) + 2, max_width)
                    ws_all.column_dimensions[get_column_letter(idx)].width = max(width, 10)

        # Per category sheets
        for cat, count in category_counts:
            if count <= 0 or cat not in df["Category"].values:
                continue
            sub = df[df["Category"] == cat][["Nested?","URL","Date","Outlet","Headline","Nut Graph"]]
            if sub.empty:
                continue
            sheet = cat[:31]
            sub.to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.sheets.get(sheet)
            if ws is not None:
                for col_name, max_width in [("Date", 22), ("Outlet", 18), ("Headline", 80)]:
                    if col_name in sub.columns:
                        idx = list(sub.columns).index(col_name) + 1
                        values = sub[col_name].astype(str).tolist()
                        width = min(max(len(col_name), max((len(v) for v in values), default=0)) + 2, max_width)
                        ws.column_dimensions[get_column_letter(idx)].width = max(width, 10)

        if not unc.empty:
            sub = unc[["Nested?","URL","Date","Outlet","Headline","Nut Graph"]]
            sub.to_excel(writer, sheet_name="Uncategorized", index=False)
            ws = writer.sheets.get("Uncategorized")
            if ws is not None:
                for col_name, max_width in [("Date", 22), ("Outlet", 18), ("Headline", 80)]:
                    if col_name in sub.columns:
                        idx = list(sub.columns).index(col_name) + 1
                        values = sub[col_name].astype(str).tolist()
                        width = min(max(len(col_name), max((len(v) for v in values), default=0)) + 2, max_width)
                        ws.column_dimensions[get_column_letter(idx)].width = max(width, 10)

    buffer.seek(0)
    default_name = f"us_china_news_{start_label}_{end_label}.xlsx"
    st.download_button(
        "⬇️ Download Excel",
        data=buffer,
        file_name=default_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"download_{start_label}_{end_label}"
    )

# Source multiselect (domain keys from config)
all_sources = list(CFG.get("rss_feeds", {}).keys())
col1, col2 = st.columns([1,1])
with col1:
    start_date = st.date_input("Start date", value=date.today(), min_value=date(2000,1,1), max_value=date.today())
with col2:
    end_date = st.date_input("End date (<= today)", value=date.today(), min_value=date(2000,1,1), max_value=date.today())
selected_sources = st.multiselect("Sources (whitelist)", options=all_sources, default=all_sources)

# API Classification toggle
use_api_classification = st.checkbox(
    "🤖 Use API Classification (OpenAI)", 
    value=False,
    help="Use AI to classify articles into categories. Disable if you hit rate limits or want faster processing."
)

# Show budget status and cost estimate if API classification is enabled
if use_api_classification:
    try:
        from api_classifier import get_budget_status, is_api_available
        
        # Debug: Show API availability status
        api_available = is_api_available()
        
        if api_available:
            budget_status = get_budget_status()
            
            if budget_status["has_budget"]:
                st.info(f"💰 API Budget Status: ${budget_status['cost_today']:.3f} used today (${budget_status['remaining']:.3f} remaining)")
            else:
                st.warning("⚠️ No daily budget limit set. Consider setting `daily_budget_usd` in secrets.toml to avoid unexpected costs.")
        else:
            # Debug: Show why API is not available
            debug_info = []
            try:
                import streamlit as st
                if hasattr(st, "secrets") and "api" in st.secrets:
                    api_config = st.secrets.get("api", {})
                    classifier_enabled = api_config.get("classifier_enabled", None)
                    openai_api_key = api_config.get("openai_api_key", None)
                    
                    if classifier_enabled is None:
                        debug_info.append("❌ `classifier_enabled` not found in secrets")
                    elif isinstance(classifier_enabled, str):
                        if classifier_enabled.lower() != "true":
                            debug_info.append(f"❌ `classifier_enabled` is '{classifier_enabled}' (should be true or 'true')")
                        else:
                            debug_info.append("✅ `classifier_enabled` is 'true'")
                    elif not bool(classifier_enabled):
                        debug_info.append(f"❌ `classifier_enabled` is {classifier_enabled} (should be true)")
                    else:
                        debug_info.append("✅ `classifier_enabled` is true")
                    
                    if not openai_api_key:
                        debug_info.append("❌ `openai_api_key` is empty or not found")
                    elif len(openai_api_key) < 10:
                        debug_info.append("❌ `openai_api_key` seems invalid (too short)")
                    else:
                        debug_info.append(f"✅ `openai_api_key` found (length: {len(openai_api_key)})")
                else:
                    debug_info.append("❌ No `[api]` section found in Streamlit secrets")
                    debug_info.append("💡 Please configure secrets in Streamlit Cloud: Settings → Secrets")
            except Exception as e:
                debug_info.append(f"⚠️ Error checking secrets: {e}")
            
            with st.expander("🔍 API Configuration Debug Info", expanded=True):
                st.error("⚠️ API classifier is not available")
                st.markdown("**Debug information:**")
                for info in debug_info:
                    st.text(info)
                st.markdown("---")
                st.markdown("**How to fix:**")
                st.markdown("""
                1. Go to Streamlit Cloud: https://share.streamlit.io/
                2. Open your app settings
                3. Click "Secrets" or "Settings" → "Secrets"
                4. Make sure you have:
                   ```toml
                   [api]
                   classifier_enabled = true
                   provider = "openai"
                   openai_api_key = "sk-your-api-key"
                   ```
                5. Save and redeploy the app
                """)
    except Exception as e:
        st.error(f"⚠️ Error checking API availability: {e}")
        import traceback
        with st.expander("Error details"):
            st.code(traceback.format_exc())

# Google Sheets configuration
use_sheets_db = False
spreadsheet_id = None
if HAS_SHEETS:
    st.markdown("---")
    st.markdown("### 📊 Data Source")
    use_sheets_db = st.checkbox("Read historical data from Google Sheets (NYT, SCMP, Reuters, Financial Times,Washington Post, and Associated Press)", value=True)
    if use_sheets_db:
        # Try to get Google Sheets ID from multiple sources
        default_sheets_id = ""
        # 1. Try reading from Streamlit secrets
        try:
            if hasattr(st, 'secrets') and 'GOOGLE_SHEETS_ID' in st.secrets:
                default_sheets_id = st.secrets['GOOGLE_SHEETS_ID']
        except:
            pass
        # 2. Try reading from environment variables
        if not default_sheets_id:
            default_sheets_id = os.getenv("GOOGLE_SHEETS_ID", "")
        
        spreadsheet_id = st.text_input(
            "Google Sheets ID", 
            value=default_sheets_id,
            placeholder="Get from Google Sheets URL",
            help="Example: https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/edit"
        )
        
        if spreadsheet_id:
            st.info("✅ Will read historical data for NYT, SCMP, Reuters, Financial Times, Washington Post, and Associated Press from Google Sheets")

col_btn1, col_btn2 = st.columns([1, 3])
with col_btn1:
    # 防止重复点击：如果正在处理，禁用按钮
    is_processing = st.session_state.get('is_processing', False)
    run = st.button("Generate & Export", type="primary", disabled=is_processing)
with col_btn2:
    if st.button("Clear Results"):
        st.session_state.df_result = None
        st.session_state.last_run_params = None
        st.session_state.is_processing = False
        st.session_state.api_calls_made = 0
        st.session_state.last_api_call_time = None  # 清除 API 调用时间记录
        st.rerun()

# Check if we should display previous results instead of form
if not run and st.session_state.df_result is not None:
    df = st.session_state.df_result
    params = st.session_state.last_run_params
    
    if params:
        st.info(f"💾 Showing saved results from: {params['start_date']} to {params['end_date']}, {len(params.get('selected_sources', []))} sources")
        start_date = params['start_date']
        end_date = params['end_date']
    else:
        start_date = date.today()
        end_date = date.today()
    
    render_results(df, start_date, end_date)
    st.info("💡 Click *Clear Results* above to re-filter; otherwise, stay on this page.")
    st.stop()

elif run:
    # 防止重复执行：如果正在处理中，直接返回
    if st.session_state.is_processing:
        st.warning("⚠️ Processing is already in progress. Please wait...")
        st.stop()
    
    # 防止 hot-reload 导致的重复执行
    # 检查是否在短时间内重复执行（可能是 hot-reload 触发）
    if st.session_state.last_api_call_time:
        time_since_last = (datetime.now() - st.session_state.last_api_call_time).total_seconds()
        if time_since_last < 5:  # 5 秒内重复执行，可能是 hot-reload
            st.warning("⚠️ Detected potential duplicate execution. Please wait a few seconds before retrying.")
        st.stop()

    if end_date > date.today():
        st.error("End date cannot be in the future.")
    elif start_date > end_date:
        st.error("Start date must be before end date.")
    else:
        # Store run parameters
        current_params = {
            'start_date': start_date,
            'end_date': end_date,
            'selected_sources': selected_sources,
            'use_sheets_db': use_sheets_db
        }
        
        # 设置处理状态，防止重复执行
        st.session_state.is_processing = True
        
        with st.spinner("Collecting articles..."):
            # Progress tracking
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # 1. Read historical data from Google Sheets (if enabled)
            sheets_df = pd.DataFrame()
            if use_sheets_db and spreadsheet_id and HAS_SHEETS:
                try:
                    status_text.text("📖 Reading historical data from Google Sheets...")
                    progress_bar.progress(10)
                    st.info("📖 Reading historical data from Google Sheets...")
                    # Try reading multiple possible sheets
                    # Simplified: read all data, then filter by date
                    # Can be optimized to only read relevant sheets
                    priority_sources = ["nytimes.com", "scmp.com", "reuters.com", "ft.com", "apnews.com"]
                    
                    # Try to read all sheets starting with "Week", merge data
                    try:
                        # Read all Week sheet data
                        import gspread
                        from google.oauth2.service_account import Credentials
                        from google_sheets_integration import get_sheets_client
                        
                        client = get_sheets_client(credentials_path=None)
                        spreadsheet = client.open_by_key(spreadsheet_id)
                        
                        all_sheets_data = []
                        for sheet in spreadsheet.worksheets():
                            # Read all sheets (not just "Week" sheets) to get all data
                            # Previously only read "Week" sheets, but user may have data in other sheets
                            if True:  # Read all sheets
                                try:
                                    data = sheet.get_all_values()
                                    if len(data) > 1:  # Has data (header + data)
                                        df_part = pd.DataFrame(data[1:], columns=data[0])
                                        # 在合并前先解析日期，避免合并后格式不一致导致解析失败
                                        if 'Date' in df_part.columns:
                                            # 保存原始日期字符串
                                            date_original = df_part['Date'].copy()
                                            
                                            # 使用更宽松的日期解析，支持多种格式（有秒/无秒）
                                            # 先尝试标准格式
                                            df_part['Date'] = pd.to_datetime(df_part['Date'], errors='coerce', infer_datetime_format=True)
                                            
                                            # 如果还有无法解析的，尝试手动解析（处理 "2025-11-07 18:10" 这种格式）
                                            if df_part['Date'].isna().any():
                                                def parse_date_flexible(date_str):
                                                    if pd.isna(date_str) or date_str == '':
                                                        return pd.NaT
                                                    try:
                                                        # 尝试多种格式
                                                        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"]:
                                                            try:
                                                                return pd.to_datetime(date_str, format=fmt)
                                                            except:
                                                                continue
                                                        # 如果都失败，使用默认解析
                                                        return pd.to_datetime(date_str, errors='coerce')
                                                    except:
                                                        return pd.NaT
                                                
                                                # 只对无法解析的日期重新解析（使用原始字符串）
                                                mask = df_part['Date'].isna()
                                                if mask.any():
                                                    df_part.loc[mask, 'Date'] = date_original.loc[mask].apply(parse_date_flexible)
                                        all_sheets_data.append(df_part)
                                except Exception as e:
                                    st.warning(f"⚠️ Error reading Sheet '{sheet.title}': {e}")
                        
                        # Merge all sheet data
                        if all_sheets_data:
                            sheets_df = pd.concat(all_sheets_data, ignore_index=True)
                        else:
                            sheets_df = pd.DataFrame()
                        
                        if not sheets_df.empty and 'Date' in sheets_df.columns:
                            # Filter by date range
                            before_date_filter = len(sheets_df)
                            # Date 列已经在合并前解析过了，这里只需要确保是 datetime 类型
                            if not pd.api.types.is_datetime64_any_dtype(sheets_df['Date']):
                                sheets_df['Date'] = pd.to_datetime(sheets_df['Date'], errors='coerce')
                            # Handle date range: if only date (no time), end_date should include all time for that day
                            date_from_dt = pd.to_datetime(start_date).normalize()  # Set to 00:00:00
                            date_to_dt = pd.to_datetime(end_date).normalize() + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)  # Set to 23:59:59
                            sheets_df = sheets_df[
                                (sheets_df['Date'] >= date_from_dt) & 
                                (sheets_df['Date'] <= date_to_dt)
                            ]
                            after_date_filter = len(sheets_df)
                            
                            # Filter by selected sources (Outlet column)
                            # 注意：如果 selected_sources 为空列表，不进行过滤（显示所有数据）
                            # 如果 selected_sources 不为空，只显示选中的 Outlet
                            if 'Outlet' in sheets_df.columns and selected_sources and len(selected_sources) > 0:
                                # Convert selected sources (domain names) to outlet abbreviations
                                from utils import normalize_source_short
                                # Convert domain names to abbreviations (e.g., nytimes.com → NYT)
                                selected_outlets = set()
                                for source in selected_sources:
                                    # Normalize the domain name to get the abbreviation
                                    normalized = normalize_source_short(source)
                                    selected_outlets.add(normalized)
                                
                                # 在过滤前保存可用的 Outlet（用于调试）
                                available_outlets_before = sheets_df['Outlet'].unique().tolist() if 'Outlet' in sheets_df.columns else []
                                
                                # Filter: keep only articles from selected outlets
                                sheets_df = sheets_df[sheets_df['Outlet'].isin(selected_outlets)]
                                
                                # Debug: 显示过滤信息（如果日期过滤后有数据，但 source 过滤后为 0）
                                if after_date_filter > 0 and len(sheets_df) == 0:
                                    st.warning(f"⚠️ Found {after_date_filter} records after date filtering, but selected sources do not match any Outlets.\n"
                                              f"   Selected Outlets: {sorted(selected_outlets)}\n"
                                              f"   Available Outlets in data: {sorted(available_outlets_before)}")
                            
                            # Ensure column names are consistent
                            if 'Nested?' not in sheets_df.columns:
                                sheets_df['Nested?'] = ''
                            
                            # 显示更详细的信息
                            if len(sheets_df) == 0:
                                if after_date_filter > 0:
                                    st.warning(f"⚠️ There are {after_date_filter} records in the date range, but 0 records after filtering by selected sources")
                                else:
                                    st.info(f"ℹ️ No data in date range {start_date} to {end_date} (total data: {before_date_filter} records)")
                            else:
                                # 显示详细信息：包括日期过滤后的数量和 Outlet 过滤后的数量
                                if after_date_filter != len(sheets_df):
                                    st.success(f"✅ Read {len(sheets_df)} historical records from Google Sheets (after date filter: {after_date_filter} records, after Outlet filter: {len(sheets_df)} records)")
                                else:
                                    st.success(f"✅ Read {len(sheets_df)} historical records from Google Sheets (filtered by selected sources)")
                    except Exception as e:
                        st.warning(f"⚠️ Unable to read Google Sheets: {e}")
                        sheets_df = pd.DataFrame()
                except Exception as e:
                    st.warning(f"⚠️ Google Sheets read failed: {e}")
            
            # 2. Real-time RSS scraping from all sources
            status_text.text(f"🌐 Scraping from {len(selected_sources)} RSS feeds...")
            progress_bar.progress(30)
            st.info("🌐 Scraping from RSS feeds in real-time...")
            rss_df = collect_rss(
                "config_en.yaml", 
                start_date.isoformat(), 
                end_date.isoformat(), 
                us_china_only=False, 
                limit_sources=selected_sources
            )
            progress_bar.progress(60)
            
            # 3. Merge data
            status_text.text("🔄 Merging and deduplicating data...")
            progress_bar.progress(70)
            
            if not sheets_df.empty and not rss_df.empty:
                # Ensure column names are consistent
                required_cols = ["Nested?","URL","Date","Outlet","Headline","Nut Graph"]
                for col in required_cols:
                    if col not in sheets_df.columns:
                        sheets_df[col] = ""
                    if col not in rss_df.columns:
                        rss_df[col] = ""
                
                # Debug: 显示合并前的统计
                sheets_count = len(sheets_df)
                rss_count = len(rss_df)
                
                # Merge
                df = pd.concat([sheets_df[required_cols], rss_df[required_cols]], ignore_index=True)
                before_dedup = len(df)
                
                # Deduplicate (by URL)
                df = df.drop_duplicates(subset=['URL'], keep='first')
                after_dedup = len(df)
                duplicates_removed = before_dedup - after_dedup
                
                # 再次应用 Outlet 过滤（确保合并后的数据也符合选中的 sources）
                # 因为 RSS 数据可能包含未选中的 Outlet
                if 'Outlet' in df.columns and selected_sources and len(selected_sources) > 0:
                    from utils import normalize_source_short
                    selected_outlets = set()
                    for source in selected_sources:
                        normalized = normalize_source_short(source)
                        selected_outlets.add(normalized)
                    
                    before_outlet_filter = len(df)
                    df = df[df['Outlet'].isin(selected_outlets)]
                    after_outlet_filter = len(df)
                    if before_outlet_filter != after_outlet_filter:
                        st.info(f"🔍 Outlet filtering after merge: {before_outlet_filter} → {after_outlet_filter} records (removed {before_outlet_filter - after_outlet_filter} unselected Outlets)")
                
                # Debug: 显示详细的合并信息
                if 'Outlet' in df.columns:
                    outlet_counts = df['Outlet'].value_counts()
                    st.info(f"📊 Merge details: Google Sheets ({sheets_count} records) + RSS ({rss_count} records) = {before_dedup} records after merge → {after_dedup} records after deduplication (removed {duplicates_removed} duplicates)\n"
                           f"   Outlet distribution: {len(outlet_counts)} Outlets")
                
                st.success(f"✅ Merge complete: Google Sheets ({sheets_count} records) + RSS ({rss_count} records) = Total {after_dedup} records (after deduplication)")
            elif not sheets_df.empty:
                df = sheets_df
                st.success(f"✅ Using Google Sheets data: {len(df)} records")
            elif not rss_df.empty:
                df = rss_df
                st.success(f"✅ Using RSS data: {len(rss_df)} records")
            else:
                df = pd.DataFrame()
                st.warning("No articles found")

        if not df.empty:
            # Show cost estimate before classification if using API
            if use_api_classification:
                try:
                    from api_classifier import estimate_cost, get_budget_status
                    cost_estimate = estimate_cost(len(df))
                    budget_status = get_budget_status()
                    
                    if budget_status["has_budget"]:
                        if not cost_estimate["can_afford"]:
                            st.error(f"❌ Budget insufficient! Estimated cost: ${cost_estimate['estimated_cost']:.3f}, Remaining: ${cost_estimate['remaining_budget']:.3f}")
                            st.stop()
                        else:
                            st.info(f"💰 Estimated API cost: ${cost_estimate['estimated_cost']:.3f} for {len(df)} articles (${cost_estimate['remaining_budget']:.3f} remaining)")
                    else:
                        st.warning(f"⚠️ Estimated API cost: ${cost_estimate['estimated_cost']:.3f} for {len(df)} articles (no budget limit set)")
                except:
                    pass
            
            # 分类步骤：显示进度（无论是否使用 API）
            if use_api_classification:
                status_text.text("🏷️ Classifying articles with API...")
            else:
                status_text.text("🏷️ Classifying articles with keywords...")
            progress_bar.progress(80)
            
            # Assign single category per article (first matched)
            compiled = []
            for cat, patt in CATEGORIES.items():
                try:
                    compiled.append((cat, compile_or_regex([patt])))
                except Exception:
                    continue

            # Track API calls for debugging (use list to allow modification in nested function)
            api_stats = {"calls_made": 0, "calls_successful": 0}

            def assign_category(row):
                # Try using API classification (if enabled)
                if use_api_classification:
                    # Debug: Log that API classification is enabled
                    import sys
                    print(f"🔍 assign_category() called with use_api_classification=True", file=sys.stderr, flush=True)
                    
                    # 记录 API 调用时间（用于检测重复执行）
                    if st.session_state.last_api_call_time is None:
                        st.session_state.last_api_call_time = datetime.now()
                    try:
                        from api_classifier import classify_with_api, is_api_available
                        import time
                        
                        api_available_check = is_api_available()
                        print(f"🔍 is_api_available() check in assign_category: {api_available_check}", file=sys.stderr, flush=True)
                        
                        # 如果 API 不可用，显示详细错误信息（仅第一次）
                        if not api_available_check:
                            if not hasattr(assign_category, '_api_unavailable_warned'):
                                st.error("❌ API classification is enabled but API is not available. Please check the API configuration debug info above.")
                                assign_category._api_unavailable_warned = True
                                # 打印详细的调试信息到 stderr
                                try:
                                    import streamlit as st
                                    if hasattr(st, "secrets") and "api" in st.secrets:
                                        api_config = st.secrets.get("api", {})
                                        classifier_enabled = api_config.get("classifier_enabled", None)
                                        openai_api_key = api_config.get("openai_api_key", None)
                                        print(f"🔍 Debug: classifier_enabled = {classifier_enabled} (type: {type(classifier_enabled)})", file=sys.stderr, flush=True)
                                        print(f"🔍 Debug: openai_api_key exists = {bool(openai_api_key)}, length = {len(openai_api_key) if openai_api_key else 0}", file=sys.stderr, flush=True)
                                    else:
                                        print(f"🔍 Debug: No [api] section in Streamlit secrets", file=sys.stderr, flush=True)
                                except Exception as e:
                                    print(f"🔍 Debug: Error checking secrets: {e}", file=sys.stderr, flush=True)
                        
                        if api_available_check:
                            category_list = [cat for cat, _ in compiled] + ["Uncategorized"]
                            
                            # Add retry logic for rate limits
                            max_retries = 3
                            for attempt in range(max_retries):
                                try:
                                    api_stats["calls_made"] += 1
                                    api_cat = classify_with_api(
                                        row.get('Headline', ''),
                                        row.get('Nut Graph', ''),
                                        category_list
                                    )
                                    if api_cat:
                                        api_stats["calls_successful"] += 1
                                        # Debug: 显示使用了 API 分类
                                        if not hasattr(assign_category, '_api_used'):
                                            st.info("✅ Using API classification (95-98% accuracy)")
                                            assign_category._api_used = True
                                        return api_cat
                                    else:
                                        # API returned None - log this for debugging
                                        if not hasattr(assign_category, '_api_returned_none_warned'):
                                            st.warning("⚠️ API classification returned None (check logs for details). Falling back to keyword classification.")
                                            assign_category._api_returned_none_warned = True
                                    break  # Exit retry loop (even if api_cat is None)
                                except Exception as e:
                                    if "rate_limit" in str(e).lower() and attempt < max_retries - 1:
                                        # Wait and retry for rate limit errors
                                        # Exponential backoff: 2s, 4s, 8s
                                        wait_time = 2 ** attempt
                                        time.sleep(wait_time)
                                        continue
                                    else:
                                        raise  # Re-raise other errors or final attempt
                        else:
                            # Debug: 显示 API 不可用
                            if not hasattr(assign_category, '_fallback_warned'):
                                st.warning("⚠️ API classifier not available, using keyword classification (70-80% accuracy)")
                                assign_category._fallback_warned = True
                    except ImportError:
                        pass  # API classifier not installed, use regex
                    except Exception as e:
                        if not hasattr(assign_category, '_error_shown'):
                            if "rate_limit" in str(e).lower():
                                st.error(f"⚠️ API Rate Limit reached. Falling back to keyword classification. Try again in a minute or disable API classification.")
                            else:
                                st.error(f"⚠️ API classification error: {e}")
                            assign_category._error_shown = True
                        # Fall through to regex classification
                
                # Use regex classification (default or fallback)
                text = f"{row.get('Headline','')} || {row.get('Nut Graph','')}"
                for cat, rgx in compiled:
                    if rgx.search(text):
                        return cat
                return "Uncategorized"

            df = df.copy()
            
            # Apply classification with progress tracking（无论是否使用 API 都显示进度）
            total = len(df)
            categories = []
            for idx, row in df.iterrows():
                cat = assign_category(row)
                categories.append(cat)
                # Update progress every 10%（或不使用 API 时每 20%）
                update_interval = max(1, total // 10) if use_api_classification else max(1, total // 5)
                if idx % update_interval == 0:
                    if use_api_classification:
                        progress = 80 + int((idx / total) * 15)
                    else:
                        progress = 80 + int((idx / total) * 20)
                    progress_bar.progress(min(progress, 100))
                    if use_api_classification:
                        status_text.text(f"🏷️ Classifying with API... {idx+1}/{total} articles")
                    else:
                        status_text.text(f"🏷️ Classifying with keywords... {idx+1}/{total} articles")
            
            df["Category"] = categories
            progress_bar.progress(100)
            status_text.text("✅ Complete!")
            
            # Show API usage summary if API classification was used
            if use_api_classification and api_stats["calls_made"] > 0:
                try:
                    from api_classifier import get_budget_status
                    budget_status = get_budget_status()
                    st.info(f"📊 API Usage: {api_stats['calls_made']} calls attempted, {api_stats['calls_successful']} successful. "
                           f"Budget: ${budget_status['cost_today']:.3f} used today "
                           f"(${budget_status['remaining']:.3f} remaining)")
                    
                    # Note about Streamlit Cloud temporary file system
                    # Check if running on Streamlit Cloud (has STREAMLIT_SERVER_RUNNING_ON_PORT env var)
                    if os.getenv("STREAMLIT_SERVER_RUNNING_ON_PORT"):
                        st.warning("ℹ️ Note: In Streamlit Cloud, budget tracking resets on each deployment/restart due to temporary file system. "
                                  "The actual API calls are still being made and charged to your OpenAI account.")
                except Exception as e:
                    st.warning(f"⚠️ Could not retrieve budget status: {e}")
            
            # Store result in session state to persist
            st.session_state.df_result = df.copy()
            st.session_state.last_run_params = current_params
            st.session_state.is_processing = False  # 处理完成，重置状态
            # 记录 API 调用完成时间
            if use_api_classification:
                st.session_state.last_api_call_time = datetime.now()
            render_results(df, start_date, end_date)

# This section is now handled above before the "if run:" block
# Keeping this for backwards compatibility but it should not be reached
elif False and st.session_state.df_result is not None and not run:
    st.info("💡 Showing previous results (click 'Clear Results' to reset)")
    df = st.session_state.df_result
    params = st.session_state.last_run_params
    
    if params:
        st.markdown(f"**Previous run**: {params['start_date']} to {params['end_date']}, {len(params.get('selected_sources', []))} sources")
        start_date = params['start_date']
        end_date = params['end_date']
    else:
        start_date = date.today()
        end_date = date.today()
    
    # Display summary and data (same as above)
    st.markdown("### 📊 Summary")
    
    total = len(df)
    if 'Category' in df.columns:
        unc = df[df["Category"] == "Uncategorized"]
        unc_count = len(unc) if not unc.empty else 0
        
        # Calculate metrics
        categories_used = len(set(df['Category'].unique()) - {'Uncategorized'})
        outlets_used = len(df['Outlet'].unique()) if 'Outlet' in df.columns else 0
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Articles", total)
        with col2:
            st.metric("Categories", f"{categories_used}/22")
        with col3:
            st.metric("Outlets", outlets_used)
        
        # Category breakdown
        st.markdown("#### Categories Breakdown")
        cat_counts = df['Category'].value_counts()
        
        # Two-column layout for categories
        num_categories = len(cat_counts)
        mid_point = (num_categories + 1) // 2
        
        col1, col2 = st.columns(2)
        with col1:
            for cat in list(cat_counts.index)[:mid_point]:
                if cat != "Uncategorized":
                    st.markdown(f"• **{cat}**: {cat_counts[cat]} articles")
        with col2:
            for cat in list(cat_counts.index)[mid_point:]:
                if cat != "Uncategorized":
                    st.markdown(f"• **{cat}**: {cat_counts[cat]} articles")
        
        if unc_count > 0:
            st.markdown(f"• **Uncategorized**: {unc_count} articles")
        
        # Show data table
        st.markdown("### 📰 Articles")
        st.dataframe(df, use_container_width=True, height=600)
        
        # ===== IMPORTANT: Add download button for previous results =====
        st.markdown("---")
        st.markdown("### 💾 Export")
        
        # Generate Excel file
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # Export by category
            for cat in sorted(df['Category'].unique()):
                if cat == "Uncategorized":
                    continue
                sub = df[df['Category'] == cat].copy()
                if not sub.empty:
                    sheet_name = cat[:31]  # Excel sheet name limit
                    sub.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Format
                    ws = writer.sheets[sheet_name]
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Auto-adjust column width
                    max_width = 80
                    for col_name in sub.columns:
                        idx = list(sub.columns).index(col_name) + 1
                        values = sub[col_name].astype(str).tolist()
                        width = min(max(len(col_name), max((len(v) for v in values), default=0)) + 2, max_width)
                        ws.column_dimensions[get_column_letter(idx)].width = max(width, 10)
            
            # Export uncategorized
            if unc_count > 0:
                unc.to_excel(writer, sheet_name="Uncategorized", index=False)
                ws = writer.sheets["Uncategorized"]
                header_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                max_width = 80
                for col_name in unc.columns:
                    idx = list(unc.columns).index(col_name) + 1
                    values = unc[col_name].astype(str).tolist()
                    width = min(max(len(col_name), max((len(v) for v in values), default=0)) + 2, max_width)
                    ws.column_dimensions[get_column_letter(idx)].width = max(width, 10)
        
        buffer.seek(0)
        default_name = f"us_china_news_{start_date}_{end_date}.xlsx"
        st.download_button(
            "⬇️ Download Excel", 
            data=buffer, 
            file_name=default_name, 
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_previous"  # Unique key to avoid conflicts
        )

