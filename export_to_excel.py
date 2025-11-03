import argparse, yaml, pandas as pd
from openpyxl.utils import get_column_letter
from collector import collect
from utils import compile_or_regex

def export(config_path: str, categories_path: str, out_path: str,
           date_from: str, date_to: str, us_china_only: bool,
           selected_sources: list[str] | None = None):

    df = collect(config_path, date_from, date_to, us_china_only, selected_sources)
    with open(categories_path, "r", encoding="utf-8") as f:
        cats = yaml.safe_load(f) or {}
    rules: dict = cats.get("categories", {})

    compiled = [(cat, compile_or_regex([patt])) for cat, patt in rules.items()]

    def assign_category(row):
        # 尝试使用 API 分类（如果启用）
        try:
            from api_classifier import classify_with_api, is_api_available
            if is_api_available():
                category_list = [cat for cat, _ in compiled] + ["Uncategorized"]
                api_cat = classify_with_api(
                    row.get('Headline', ''),
                    row.get('Nut Graph', ''),
                    category_list
                )
                if api_cat:
                    return api_cat
        except ImportError:
            pass  # API 分类器未安装，使用正则
        
        # 使用正则表达式分类（默认）
        text = f"{row.get('Headline','')} || {row.get('Nut Graph','')}"
        for cat, rgx in compiled:
            if rgx.search(text):
                return cat
        return "Uncategorized"

    df = df.copy()
    df["Category"] = df.apply(assign_category, axis=1)
    df = df.sort_values("Date", ascending=False).drop_duplicates(subset=["URL"], keep="first")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        all_cols = ["Nested?","URL","Date","Outlet","Headline","Nut Graph"]
        df[all_cols].to_excel(writer, sheet_name="All", index=False)
        # Autofit All
        ws_all = writer.sheets.get("All")
        if ws_all is not None:
            for col_name, max_width in [("Date", 22), ("Outlet", 18), ("Headline", 80)]:
                if col_name in all_cols:
                    idx = all_cols.index(col_name) + 1
                    values = df[col_name].astype(str).tolist() if not df.empty else []
                    width = min(max(len(col_name), max((len(v) for v in values), default=0)) + 2, max_width)
                    ws_all.column_dimensions[get_column_letter(idx)].width = max(width, 10)
        for cat, _ in compiled:
            sub = df[df["Category"] == cat]
            if not sub.empty:
                sheet = cat[:31]
                sub[all_cols].to_excel(writer, sheet_name=sheet, index=False)
                ws = writer.sheets.get(sheet)
                if ws is not None:
                    for col_name, max_width in [("Date", 22), ("Outlet", 18), ("Headline", 80)]:
                        idx = all_cols.index(col_name) + 1
                        values = sub[col_name].astype(str).tolist()
                        width = min(max(len(col_name), max((len(v) for v in values), default=0)) + 2, max_width)
                        ws.column_dimensions[get_column_letter(idx)].width = max(width, 10)
        unc = df[df["Category"] == "Uncategorized"]
        if not unc.empty:
            sheet = "Uncategorized"
            unc[all_cols].to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.sheets.get(sheet)
            if ws is not None:
                for col_name, max_width in [("Date", 22), ("Outlet", 18), ("Headline", 80)]:
                    idx = all_cols.index(col_name) + 1
                    values = unc[col_name].astype(str).tolist()
                    width = min(max(len(col_name), max((len(v) for v in values), default=0)) + 2, max_width)
                    ws.column_dimensions[get_column_letter(idx)].width = max(width, 10)

    print(f"Excel written: {out_path}")

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", default="config_en.yaml")
    ap.add_argument("--categories", default="categories_en.yaml")
    ap.add_argument("--out", required=True)
    ap.add_argument("--date_from", required=True)
    ap.add_argument("--date_to", required=True)
    ap.add_argument("--us_china_only", action="store_true", default=False)
    ap.add_argument("--sources", nargs="*", default=None)
    args = ap.parse_args()

    export(args.config, args.categories, args.out,
           args.date_from, args.date_to, args.us_china_only, args.sources)