import streamlit as st
from datetime import date, datetime
import pandas as pd
from collector import collect
from export_to_excel import export
import yaml, io
from utils import compile_or_regex
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="U.S.-China News Scraper", layout="wide")

st.markdown("## U.S.-China News Scraper")

# Load config and categories
with open("config_en.yaml","r",encoding="utf-8") as f:
    CFG = yaml.safe_load(f) or {}
with open("categories_en.yaml","r",encoding="utf-8") as f:
    CATS = yaml.safe_load(f) or {}
CATEGORIES = CATS.get("categories", {})

# Source multiselect (domain keys from config)
all_sources = list(CFG.get("rss_feeds", {}).keys())
col1, col2 = st.columns([1,1])
with col1:
    start_date = st.date_input("Start date", value=date.today())
with col2:
    end_date = st.date_input("End date (<= today)", value=date.today(), min_value=date(2000,1,1), max_value=date.today())
selected_sources = st.multiselect("Sources (whitelist)", options=all_sources, default=all_sources)

run = st.button("Generate Excel")

if run:
    if end_date > date.today():
        st.error("End date cannot be in the future.")
    elif start_date > end_date:
        st.error("Start date must be before end date.")
    else:
        with st.spinner("Collecting..."):
            df = collect("config_en.yaml", start_date.isoformat(), end_date.isoformat(), us_china_only=False, limit_sources=selected_sources)

        st.success(f"Collected {len(df)} articles.")

        # Assign single category per article (first matched)
        compiled = []
        for cat, patt in CATEGORIES.items():
            try:
                compiled.append((cat, compile_or_regex([patt])))
            except Exception:
                continue

        def assign_category(row):
            # 尝试使用 API 分类（如果启用）
            try:
                from api_classifier import classify_with_api, is_api_available
                if is_api_available():
                    category_list = [cat for cat, _ in compiled] + ["Uncategorized"]
                    api_cat = classify_with_api(
                        row.get('Headline', ''),
                        row.get('Nut Graph', ''),
                        category_list
                    )
                    if api_cat:
                        return api_cat
            except ImportError:
                pass  # API 分类器未安装，使用正则
            
            # 使用正则表达式分类（默认）
            text = f"{row.get('Headline','')} || {row.get('Nut Graph','')}"
            for cat, rgx in compiled:
                if rgx.search(text):
                    return cat
            return "Uncategorized"

        df = df.copy()
        df["Category"] = df.apply(assign_category, axis=1)

        # Per-category counts
        st.markdown("### Summary")
        st.write(f"- Total: **{len(df)}**")
        for cat, _ in compiled:
            sub = df[df["Category"] == cat]
            st.write(f"- {cat}: **{len(sub)}**")
        unc = df[df["Category"] == "Uncategorized"]
        if not unc.empty:
            st.write(f"- Uncategorized: **{len(unc)}**")

        # Build Excel in-memory
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # All first
            df.to_excel(writer, sheet_name="All", index=False)
            # Autofit for All
            ws_all = writer.sheets.get("All")
            if ws_all is not None:
                for col_name, max_width in [("Date", 22), ("Outlet", 18), ("Headline", 80)]:
                    if col_name in df.columns:
                        idx = list(df.columns).index(col_name) + 1
                        values = df[col_name].astype(str).tolist() if not df.empty else []
                        width = min(max(len(col_name), max((len(v) for v in values), default=0)) + 2, max_width)
                        ws_all.column_dimensions[get_column_letter(idx)].width = max(width, 10)
            # Categories (single-category assignment)
            for cat, _ in compiled:
                sub = df[df["Category"] == cat]
                if not sub.empty:
                    sub = sub[["Nested?","URL","Date","Outlet","Headline","Nut Graph"]]
                    sheet = cat[:31]
                    sub.to_excel(writer, sheet_name=sheet, index=False)
                    ws = writer.sheets.get(sheet)
                    if ws is not None:
                        for col_name, max_width in [("Date", 22), ("Outlet", 18), ("Headline", 80)]:
                            if col_name in sub.columns:
                                idx = list(sub.columns).index(col_name) + 1
                                values = sub[col_name].astype(str).tolist()
                                width = min(max(len(col_name), max((len(v) for v in values), default=0)) + 2, max_width)
                                ws.column_dimensions[get_column_letter(idx)].width = max(width, 10)
            # Uncategorized
            if not unc.empty:
                sub = unc[["Nested?","URL","Date","Outlet","Headline","Nut Graph"]]
                sub.to_excel(writer, sheet_name="Uncategorized", index=False)
                ws = writer.sheets.get("Uncategorized")
                if ws is not None:
                    for col_name, max_width in [("Date", 22), ("Outlet", 18), ("Headline", 80)]:
                        if col_name in sub.columns:
                            idx = list(sub.columns).index(col_name) + 1
                            values = sub[col_name].astype(str).tolist()
                            width = min(max(len(col_name), max((len(v) for v in values), default=0)) + 2, max_width)
                            ws.column_dimensions[get_column_letter(idx)].width = max(width, 10)

        buffer.seek(0)
        default_name = f"us_china_news_{start_date}_{end_date}.xlsx"
        st.download_button("⬇️ Download Excel", data=buffer, file_name=default_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
